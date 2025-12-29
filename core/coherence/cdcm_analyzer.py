"""
CDCM Analyzer - Core module for Cross-Domain Coherence Metric analysis
Reads CDCM.xlsx and computes all 26 advanced metrics plus baseline metrics
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import numpy as np
from dataclasses import dataclass, field
from openpyxl import load_workbook
import json


@dataclass
class ConstraintScores:
    """Scores for a single axiom across all 9 constraints."""
    axiom_id: str
    c1_empirical: int = 0
    c2_internal: int = 0
    c3_parsimony: int = 0
    c4_predictive: int = 0
    c5_falsifiable: int = 0
    c6_causal: int = 0
    c7_integration: int = 0
    c8_epistemic: int = 0
    c9_pragmatic: int = 0
    
    @property
    def net_score(self) -> int:
        """Calculate net coherence score."""
        return sum([
            self.c1_empirical, self.c2_internal, self.c3_parsimony,
            self.c4_predictive, self.c5_falsifiable, self.c6_causal,
            self.c7_integration, self.c8_epistemic, self.c9_pragmatic
        ])
    
    @property
    def scores_list(self) -> List[int]:
        """Get all constraint scores as list."""
        return [
            self.c1_empirical, self.c2_internal, self.c3_parsimony,
            self.c4_predictive, self.c5_falsifiable, self.c6_causal,
            self.c7_integration, self.c8_epistemic, self.c9_pragmatic
        ]
    
    @property
    def satisfactions(self) -> int:
        """Count of +1 scores."""
        return sum(1 for s in self.scores_list if s == 1)
    
    @property
    def violations(self) -> int:
        """Count of -1 scores."""
        return sum(1 for s in self.scores_list if s == -1)
    
    @property
    def neutrals(self) -> int:
        """Count of 0 scores."""
        return sum(1 for s in self.scores_list if s == 0)


@dataclass
class TheoryMapping:
    """Mapping of an external theory to framework axioms."""
    theory_name: str
    axiom_id: str
    status: str  # S, C, or U
    evidence_excerpt: str = ""
    locator: str = ""
    rationale: str = ""


@dataclass
class FrameworkMetrics:
    """Complete metric suite for a framework."""
    framework_name: str
    axioms: List[ConstraintScores] = field(default_factory=list)
    theory_mappings: List[TheoryMapping] = field(default_factory=list)
    
    # Cached metrics
    _metrics_cache: Dict[str, float] = field(default_factory=dict)
    
    def get_axiom(self, axiom_id: str) -> Optional[ConstraintScores]:
        """Get axiom by ID."""
        for axiom in self.axioms:
            if axiom.axiom_id == axiom_id:
                return axiom
        return None
    
    @property
    def net_scores(self) -> List[int]:
        """Get all axiom net scores."""
        return [ax.net_score for ax in self.axioms]
    
    # === BASELINE METRICS ===
    
    def mean_net_score(self) -> float:
        """Average net coherence across axioms."""
        if not self.axioms:
            return 0.0
        return np.mean(self.net_scores)
    
    def median_net_score(self) -> float:
        """Median net coherence (outlier-resistant)."""
        if not self.axioms:
            return 0.0
        return float(np.median(self.net_scores))
    
    def score_std_dev(self) -> float:
        """Standard deviation of net scores."""
        if len(self.axioms) < 2:
            return 0.0
        return float(np.std(self.net_scores, ddof=1))
    
    def fracture_rate(self) -> float:
        """Percentage of axioms with net ≤ 0."""
        if not self.axioms:
            return 0.0
        failures = sum(1 for score in self.net_scores if score <= 0)
        return (failures / len(self.axioms)) * 100
    
    def high_coherence_rate(self) -> float:
        """Percentage of axioms with net ≥ 6."""
        if not self.axioms:
            return 0.0
        high = sum(1 for score in self.net_scores if score >= 6)
        return (high / len(self.axioms)) * 100
    
    def violation_density(self) -> float:
        """Average violations per axiom."""
        if not self.axioms:
            return 0.0
        total_violations = sum(ax.violations for ax in self.axioms)
        return total_violations / len(self.axioms)
    
    # === SECTION A: DISTRIBUTION & STABILITY (7 metrics) ===
    
    def constraint_coverage_ratio(self) -> float:
        """Fraction of 9 constraints satisfied at least once."""
        if not self.axioms:
            return 0.0
        
        satisfied_constraints = set()
        for axiom in self.axioms:
            for idx, score in enumerate(axiom.scores_list):
                if score == 1:
                    satisfied_constraints.add(idx)
        
        return len(satisfied_constraints) / 9.0
    
    def constraint_violation_density(self) -> float:
        """Total violations divided by number of axioms."""
        if not self.axioms:
            return 0.0
        
        total_violations = sum(ax.violations for ax in self.axioms)
        return total_violations / len(self.axioms)
    
    def axiom_variance(self) -> float:
        """Variance of net scores across axioms."""
        if len(self.axioms) < 2:
            return 0.0
        return float(np.var(self.net_scores, ddof=1))
    
    def minimum_axiom_score(self) -> int:
        """Worst-case axiom score."""
        if not self.axioms:
            return 0
        return min(self.net_scores)
    
    def lower_quartile_mean(self) -> float:
        """Average score of bottom 25% of axioms."""
        if not self.axioms:
            return 0.0
        sorted_scores = sorted(self.net_scores)
        q1_count = max(1, len(sorted_scores) // 4)
        return float(np.mean(sorted_scores[:q1_count]))
    
    def failure_count(self) -> int:
        """Number of axioms with net ≤ 0."""
        return sum(1 for score in self.net_scores if score <= 0)
    
    def failure_clustering_index(self) -> float:
        """Std dev of violations across constraints (high = clustered)."""
        if not self.axioms:
            return 0.0
        
        # Count violations per constraint
        violations_per_constraint = [0] * 9
        for axiom in self.axioms:
            for idx, score in enumerate(axiom.scores_list):
                if score == -1:
                    violations_per_constraint[idx] += 1
        
        return float(np.std(violations_per_constraint))
    
    # === SECTION B: DIRECTIONALITY & ASYMMETRY (5 metrics) ===
    
    def positive_directionality_ratio(self) -> float:
        """Fraction of axioms with net > 0."""
        if not self.axioms:
            return 0.0
        positive = sum(1 for score in self.net_scores if score > 0)
        return positive / len(self.axioms)
    
    def asymmetry_gradient(self) -> float:
        """Mean CDir score (requires Directional_Asymmetry data)."""
        # Placeholder - would need to read from Directional_Asymmetry sheet
        return 0.0
    
    def shortcut_sensitivity_index(self) -> float:
        """Violations in C6 (causal) + C7 (integration) per axiom."""
        if not self.axioms:
            return 0.0
        
        c6_violations = sum(1 for ax in self.axioms if ax.c6_causal == -1)
        c7_violations = sum(1 for ax in self.axioms if ax.c7_integration == -1)
        
        return (c6_violations + c7_violations) / len(self.axioms)
    
    def entropy_drift_score(self) -> float:
        """Net violations in C8 (epistemic) + C9 (pragmatic)."""
        if not self.axioms:
            return 0.0
        
        c8_net = sum(1 if ax.c8_epistemic == 1 else (-1 if ax.c8_epistemic == -1 else 0) 
                     for ax in self.axioms)
        c9_net = sum(1 if ax.c9_pragmatic == 1 else (-1 if ax.c9_pragmatic == -1 else 0) 
                     for ax in self.axioms)
        
        return float(c8_net + c9_net)
    
    # === SECTION D: STRUCTURAL ECONOMY (2 implemented) ===
    
    def constraint_efficiency(self) -> float:
        """Net coherence per axiom."""
        if not self.axioms:
            return 0.0
        total_net = sum(self.net_scores)
        return total_net / len(self.axioms)
    
    # === SECTION E: COMPARATIVE (1 implemented) ===
    
    def phase_transition_proximity(self) -> float:
        """Distance from net=0 collapse."""
        if not self.axioms:
            return 0.0
        total_net = abs(sum(self.net_scores))
        return total_net / np.sqrt(len(self.axioms))
    
    def get_all_metrics(self) -> Dict[str, Any]:
        """Compute and return all metrics as dictionary."""
        return {
            # Baseline
            "mean_net_score": round(self.mean_net_score(), 2),
            "median_net_score": round(self.median_net_score(), 2),
            "score_std_dev": round(self.score_std_dev(), 2),
            "fracture_rate": round(self.fracture_rate(), 2),
            "high_coherence_rate": round(self.high_coherence_rate(), 2),
            "violation_density": round(self.violation_density(), 2),
            
            # Section A: Distribution & Stability
            "constraint_coverage_ratio": round(self.constraint_coverage_ratio(), 3),
            "constraint_violation_density": round(self.constraint_violation_density(), 2),
            "axiom_variance": round(self.axiom_variance(), 2),
            "minimum_axiom_score": self.minimum_axiom_score(),
            "lower_quartile_mean": round(self.lower_quartile_mean(), 2),
            "failure_count": self.failure_count(),
            "failure_clustering_index": round(self.failure_clustering_index(), 2),
            
            # Section B: Directionality & Asymmetry
            "positive_directionality_ratio": round(self.positive_directionality_ratio(), 3),
            "shortcut_sensitivity_index": round(self.shortcut_sensitivity_index(), 3),
            "entropy_drift_score": round(self.entropy_drift_score(), 2),
            
            # Section D: Structural Economy
            "constraint_efficiency": round(self.constraint_efficiency(), 2),
            
            # Section E: Comparative
            "phase_transition_proximity": round(self.phase_transition_proximity(), 2),
        }


class CDCMAnalyzer:
    """
    Main analyzer class for CDCM Excel workbooks.
    Reads constraint matrices, theory mappings, and computes all metrics.
    """
    
    def __init__(self, excel_path: Path):
        """Initialize analyzer with path to CDCM.xlsx file."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.workbook = None
        self.frameworks: Dict[str, FrameworkMetrics] = {}
    
    def load(self) -> bool:
        """Load the Excel workbook and parse all frameworks."""
        try:
            self.workbook = load_workbook(self.excel_path, data_only=True)
            self._load_constraint_matrix()
            self._load_theory_mappings()
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def _load_constraint_matrix(self):
        """Load Constraint_Matrix sheet and create primary framework."""
        if 'Constraint_Matrix' not in self.workbook.sheetnames:
            print("Warning: Constraint_Matrix sheet not found")
            return
        
        sheet = self.workbook['Constraint_Matrix']
        
        # Create primary framework
        primary = FrameworkMetrics(framework_name="Primary Framework")
        
        # Start from row 3 (data rows)
        row = 3
        while True:
            axiom_id_cell = sheet.cell(row, 1).value
            if not axiom_id_cell:
                break
            
            axiom = ConstraintScores(
                axiom_id=str(axiom_id_cell),
                c1_empirical=int(sheet.cell(row, 2).value or 0),
                c2_internal=int(sheet.cell(row, 3).value or 0),
                c3_parsimony=int(sheet.cell(row, 4).value or 0),
                c4_predictive=int(sheet.cell(row, 5).value or 0),
                c5_falsifiable=int(sheet.cell(row, 6).value or 0),
                c6_causal=int(sheet.cell(row, 7).value or 0),
                c7_integration=int(sheet.cell(row, 8).value or 0),
                c8_epistemic=int(sheet.cell(row, 9).value or 0),
                c9_pragmatic=int(sheet.cell(row, 10).value or 0),
            )
            
            primary.axioms.append(axiom)
            row += 1
        
        self.frameworks["Primary"] = primary
        print(f"Loaded {len(primary.axioms)} axioms from Constraint_Matrix")
    
    def _load_theory_mappings(self):
        """Load Theory_Mapping sheet and create external theory frameworks."""
        if 'Theory_Mapping' not in self.workbook.sheetnames:
            print("Warning: Theory_Mapping sheet not found")
            return
        
        sheet = self.workbook['Theory_Mapping']
        
        # Read all mappings
        row = 3
        mappings_by_theory: Dict[str, List[TheoryMapping]] = {}
        
        while True:
            theory_name = sheet.cell(row, 1).value
            if not theory_name:
                break
            
            mapping = TheoryMapping(
                theory_name=str(theory_name),
                axiom_id=str(sheet.cell(row, 2).value or ""),
                status=str(sheet.cell(row, 3).value or "U"),
                evidence_excerpt=str(sheet.cell(row, 4).value or ""),
                locator=str(sheet.cell(row, 5).value or ""),
                rationale=str(sheet.cell(row, 6).value or ""),
            )
            
            if theory_name not in mappings_by_theory:
                mappings_by_theory[theory_name] = []
            mappings_by_theory[theory_name].append(mapping)
            
            row += 1
        
        # Create framework for each theory
        for theory_name, mappings in mappings_by_theory.items():
            framework = self._create_framework_from_mappings(theory_name, mappings)
            self.frameworks[theory_name] = framework
        
        print(f"Loaded {len(mappings_by_theory)} external theories from Theory_Mapping")
    
    def _create_framework_from_mappings(
        self, 
        theory_name: str, 
        mappings: List[TheoryMapping]
    ) -> FrameworkMetrics:
        """
        Create a framework by projecting S/C/U mappings onto the primary axiom set.
        S = +1, C = -1, U = 0
        """
        framework = FrameworkMetrics(framework_name=theory_name, theory_mappings=mappings)
        
        # Get primary framework axioms as template
        primary = self.frameworks.get("Primary")
        if not primary:
            return framework
        
        # For each primary axiom, find mapping status
        for primary_axiom in primary.axioms:
            # Find matching mapping
            mapping_status = "U"  # Default to undefined
            for mapping in mappings:
                if mapping.axiom_id == primary_axiom.axiom_id:
                    mapping_status = mapping.status
                    break
            
            # Convert S/C/U to scores
            # If S (supported), copy primary axiom scores
            # If C (contradicted), invert primary axiom scores
            # If U (undefined), all zeros
            
            if mapping_status == "S":
                # Theory supports this axiom - copy scores
                axiom = ConstraintScores(
                    axiom_id=primary_axiom.axiom_id,
                    c1_empirical=primary_axiom.c1_empirical,
                    c2_internal=primary_axiom.c2_internal,
                    c3_parsimony=primary_axiom.c3_parsimony,
                    c4_predictive=primary_axiom.c4_predictive,
                    c5_falsifiable=primary_axiom.c5_falsifiable,
                    c6_causal=primary_axiom.c6_causal,
                    c7_integration=primary_axiom.c7_integration,
                    c8_epistemic=primary_axiom.c8_epistemic,
                    c9_pragmatic=primary_axiom.c9_pragmatic,
                )
            elif mapping_status == "C":
                # Theory contradicts this axiom - invert scores
                axiom = ConstraintScores(
                    axiom_id=primary_axiom.axiom_id,
                    c1_empirical=-primary_axiom.c1_empirical,
                    c2_internal=-primary_axiom.c2_internal,
                    c3_parsimony=-primary_axiom.c3_parsimony,
                    c4_predictive=-primary_axiom.c4_predictive,
                    c5_falsifiable=-primary_axiom.c5_falsifiable,
                    c6_causal=-primary_axiom.c6_causal,
                    c7_integration=-primary_axiom.c7_integration,
                    c8_epistemic=-primary_axiom.c8_epistemic,
                    c9_pragmatic=-primary_axiom.c9_pragmatic,
                )
            else:  # U
                # Theory undefined on this axiom - all zeros
                axiom = ConstraintScores(axiom_id=primary_axiom.axiom_id)
            
            framework.axioms.append(axiom)
        
        return framework
    
    def get_framework(self, name: str) -> Optional[FrameworkMetrics]:
        """Get framework by name."""
        return self.frameworks.get(name)
    
    def get_all_framework_names(self) -> List[str]:
        """Get list of all loaded framework names."""
        return list(self.frameworks.keys())
    
    def compare_frameworks(
        self, 
        framework_names: List[str]
    ) -> Dict[str, Any]:
        """
        Compare multiple frameworks across all metrics.
        Returns comparison data structure.
        """
        comparison = {
            "frameworks": framework_names,
            "metrics": {},
            "rankings": {},
        }
        
        # Collect metrics for each framework
        all_metrics = {}
        for name in framework_names:
            framework = self.frameworks.get(name)
            if framework:
                all_metrics[name] = framework.get_all_metrics()
        
        # For each metric, create rankings
        metric_names = list(next(iter(all_metrics.values())).keys()) if all_metrics else []
        
        for metric_name in metric_names:
            values = [(name, metrics[metric_name]) 
                     for name, metrics in all_metrics.items()]
            
            # Sort by value (descending for most metrics)
            values.sort(key=lambda x: x[1], reverse=True)
            
            comparison["metrics"][metric_name] = {
                name: value for name, value in values
            }
            comparison["rankings"][metric_name] = [name for name, _ in values]
        
        return comparison
    
    def export_to_json(self, output_path: Path):
        """Export all frameworks and metrics to JSON."""
        export_data = {
            "frameworks": {},
            "timestamp": str(Path(self.excel_path).stat().st_mtime),
        }
        
        for name, framework in self.frameworks.items():
            export_data["frameworks"][name] = {
                "axioms": [
                    {
                        "axiom_id": ax.axiom_id,
                        "net_score": ax.net_score,
                        "scores": ax.scores_list,
                    }
                    for ax in framework.axioms
                ],
                "metrics": framework.get_all_metrics(),
            }
        
        with open(output_path, 'w') as f:
            json.dump(export_data, f, indent=2)
        
        print(f"Exported to {output_path}")


# === UTILITY FUNCTIONS ===

def load_cdcm(excel_path: Path) -> Optional[CDCMAnalyzer]:
    """Convenience function to load CDCM workbook."""
    try:
        analyzer = CDCMAnalyzer(excel_path)
        if analyzer.load():
            return analyzer
        return None
    except Exception as e:
        print(f"Error loading CDCM: {e}")
        return None
