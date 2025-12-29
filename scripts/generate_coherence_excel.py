"""
Generate Cross-Domain Coherence Excel Workbook
===============================================
Creates a multi-sheet Excel file for scoring axioms against 9 universal constraints.

Sheets:
1. Constraints - The 9 universal constraints with justifications
2. Domain_Manifestations - How each constraint manifests across domains
3. Axiom_Scoring - Score each axiom (+1/0/-1) against constraints
4. Axiom_Totals - Computed totals (formulas)
5. Framework_Summary - Aggregate metrics

Usage:
    python generate_coherence_excel.py
    python generate_coherence_excel.py --output "my_scoring.xlsx"
"""

import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not installed. Run: pip install openpyxl")

# The 9 Universal Constraints
CONSTRAINTS = [
    ('C1', 'Binding / Cohesion', 'Coupling strength', 'Systems fragment without it'),
    ('C2', 'Resonance', 'Constructive amplification', 'No sustained signal'),
    ('C3', 'Equilibrium', 'Stability, steady states', 'Divergence dominates'),
    ('C4', 'Temporal Persistence', 'Conservation, durability', 'Information loss'),
    ('C5', 'Positive Coupling', 'Productive gain', 'Stagnation'),
    ('C6', 'Value Conservation', 'Invariance', 'Meaning drift'),
    ('C7', 'Consistency', 'Integrity', 'Contradiction'),
    ('C8', 'Minimal Perturbation', 'Noise control', 'Signal collapse'),
    ('C9', 'Boundary Regulation', 'Constraints', 'Runaway behavior'),
]

# Domain manifestations
DOMAIN_MANIFESTATIONS = [
    ('C1', 'Strong nuclear force', 'Mutual information', 'Contract binding', 'Attachment bonds', 'Covenant'),
    ('C2', 'Constructive interference', 'Signal amplification', 'Harmonization', 'Flow state', 'Worship'),
    ('C3', 'Thermodynamic equilibrium', 'Low entropy state', 'Market clearing', 'Homeostasis', 'Shalom'),
    ('C4', 'Conservation laws', 'Error correction', 'Contract duration', 'Long-term memory', 'Faithfulness'),
    ('C5', 'Attractive forces', 'Positive feedback', 'Trade surplus', 'Prosocial behavior', 'Grace'),
    ('C6', 'Energy conservation', 'Semantic preservation', 'Store of value', 'Value stability', 'Holiness'),
    ('C7', 'Non-contradiction', 'Error-free transmission', 'Contract honor', 'Cognitive coherence', 'Truth'),
    ('C8', 'Low noise', 'Clean channel', 'Price stability', 'Emotional regulation', 'Gentleness'),
    ('C9', 'Boundary conditions', 'Channel capacity', 'Budget constraint', 'Impulse control', 'Discipline'),
]

# Sample axioms (Tier 0 Primordial)
SAMPLE_AXIOMS = [
    'P0.1 Existence',
    'P0.2 Distinction',
    'P0.3 Information Primacy',
    'P0.4 Intelligibility',
    'O1.1 Information Substrate',
    'O1.2 Self-Grounding',
    'O1.3 Information Conservation',
    'O2.1 Order Requirement',
    'O2.2 Coherence Measure',
    'O2.3 Parsimony',
]

# Colors
DARK_BG = 'FF1A1A1A'
CARD_BG = 'FF252525'
HEADER_BG = 'FF7B68EE'
POSITIVE = 'FF4ECDC4'
NEGATIVE = 'FFFF6B6B'
NEUTRAL = 'FF666666'
TEXT = 'FFE0E0E0'


def create_workbook(output_path: Path):
    """Create the full Excel workbook."""
    if not OPENPYXL_AVAILABLE:
        print("Cannot create Excel file without openpyxl")
        return

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=HEADER_BG[2:], end_color=HEADER_BG[2:], fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='333333'),
        right=Side(style='thin', color='333333'),
        top=Side(style='thin', color='333333'),
        bottom=Side(style='thin', color='333333')
    )

    # Sheet 1: Constraints
    ws1 = wb.active
    ws1.title = "Constraints"
    headers1 = ['constraint_id', 'constraint_name', 'system_property', 'necessity_justification']
    ws1.append(headers1)
    for row in CONSTRAINTS:
        ws1.append(row)

    # Format header
    for col_num, cell in enumerate(ws1[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws1.column_dimensions[get_column_letter(col_num)].width = 25

    # Sheet 2: Domain Manifestations
    ws2 = wb.create_sheet("Domain_Manifestations")
    headers2 = ['constraint_id', 'physics', 'information', 'economics', 'psychology', 'theology']
    ws2.append(headers2)
    for row in DOMAIN_MANIFESTATIONS:
        ws2.append(row)

    for col_num, cell in enumerate(ws2[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws2.column_dimensions[get_column_letter(col_num)].width = 22

    # Sheet 3: Axiom Scoring
    ws3 = wb.create_sheet("Axiom_Scoring")
    headers3 = ['axiom_id', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9']
    ws3.append(headers3)

    for axiom in SAMPLE_AXIOMS:
        ws3.append([axiom, 0, 0, 0, 0, 0, 0, 0, 0, 0])

    for col_num, cell in enumerate(ws3[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws3.column_dimensions[get_column_letter(col_num)].width = 12 if col_num > 1 else 25

    # Add data validation for -1, 0, 1
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"-1,0,1"', allow_blank=True)
    dv.error = "Only -1, 0, or 1 allowed"
    dv.errorTitle = "Invalid Entry"
    ws3.add_data_validation(dv)
    dv.add(f'B2:J{len(SAMPLE_AXIOMS) + 10}')

    # Sheet 4: Axiom Totals (with formulas)
    ws4 = wb.create_sheet("Axiom_Totals")
    headers4 = ['axiom_id', 'satisfied', 'violated', 'neutral', 'net_score']
    ws4.append(headers4)

    for i, axiom in enumerate(SAMPLE_AXIOMS, start=2):
        row_num = i
        ws4.append([
            f"=Axiom_Scoring!A{row_num}",
            f"=COUNTIF(Axiom_Scoring!B{row_num}:J{row_num},1)",
            f"=COUNTIF(Axiom_Scoring!B{row_num}:J{row_num},-1)",
            f"=COUNTIF(Axiom_Scoring!B{row_num}:J{row_num},0)",
            f"=B{i}-C{i}",
        ])

    for col_num, cell in enumerate(ws4[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws4.column_dimensions[get_column_letter(col_num)].width = 15 if col_num > 1 else 25

    # Conditional formatting for net_score
    positive_fill = PatternFill(start_color=POSITIVE[2:], end_color=POSITIVE[2:], fill_type='solid')
    negative_fill = PatternFill(start_color=NEGATIVE[2:], end_color=NEGATIVE[2:], fill_type='solid')

    ws4.conditional_formatting.add(f'E2:E{len(SAMPLE_AXIOMS)+1}',
        CellIsRule(operator='greaterThan', formula=['0'], fill=positive_fill))
    ws4.conditional_formatting.add(f'E2:E{len(SAMPLE_AXIOMS)+1}',
        CellIsRule(operator='lessThan', formula=['0'], fill=negative_fill))

    # Sheet 5: Framework Summary
    ws5 = wb.create_sheet("Framework_Summary")
    summary_data = [
        ['Metric', 'Value', 'Formula'],
        ['Total Axioms', f'=COUNTA(Axiom_Totals!A:A)-1', ''],
        ['Mean Net Score', f'=AVERAGE(Axiom_Totals!E:E)', ''],
        ['Fracture Axioms (<=0)', f'=COUNTIF(Axiom_Totals!E:E,"<=0")', ''],
        ['High Coherence (>=6)', f'=COUNTIF(Axiom_Totals!E:E,">=6")', ''],
        ['Total Satisfied', f'=SUM(Axiom_Totals!B:B)', ''],
        ['Total Violated', f'=SUM(Axiom_Totals!C:C)', ''],
        ['Coherence Ratio', f'=B6/(B6+B7)', 'satisfied / (satisfied + violated)'],
    ]

    for row in summary_data:
        ws5.append(row)

    for col_num, cell in enumerate(ws5[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws5.column_dimensions[get_column_letter(col_num)].width = 25

    # Sheet 6: Instructions
    ws6 = wb.create_sheet("Instructions")
    instructions = [
        ['CROSS-DOMAIN COHERENCE METRIC'],
        [''],
        ['HOW TO USE THIS WORKBOOK'],
        [''],
        ['1. Go to Axiom_Scoring sheet'],
        ['2. For each axiom, score each constraint:'],
        ['   +1 = constraint satisfied'],
        ['    0 = neutral'],
        ['   -1 = constraint violated'],
        [''],
        ['3. Axiom_Totals computes automatically'],
        ['4. Framework_Summary shows aggregate metrics'],
        [''],
        ['CONSTRAINT DEFINITIONS'],
        [''],
        ['C1 Binding: Does it promote cohesion?'],
        ['C2 Resonance: Does it enable amplification?'],
        ['C3 Equilibrium: Does it support stability?'],
        ['C4 Persistence: Does it conserve over time?'],
        ['C5 Positive Coupling: Does it enable productive gain?'],
        ['C6 Conservation: Does it preserve value/meaning?'],
        ['C7 Consistency: Is it non-contradictory?'],
        ['C8 Precision: Does it minimize noise?'],
        ['C9 Regulation: Does it maintain boundaries?'],
        [''],
        ['INTERPRETATION'],
        [''],
        ['Net Score > 0: More coherent than not'],
        ['Net Score < 0: More fractured than coherent'],
        ['Net Score = 0: Neutral'],
        [''],
        ['Mean Net Score across all axioms = Framework Coherence'],
    ]

    for row in instructions:
        ws6.append(row)

    ws6.column_dimensions['A'].width = 60

    # Save
    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Generate Cross-Domain Coherence Excel')
    parser.add_argument('--output', '-o',
                        default=r'O:\Theophysics_Master\TM\00_VAULT_OS\Reports\cross_domain_coherence.xlsx')
    args = parser.parse_args()

    create_workbook(Path(args.output))


if __name__ == '__main__':
    main()
